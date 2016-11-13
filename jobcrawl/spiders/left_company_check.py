import sys
import codecs
import scrapy
import locale
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

from scrapy import signals
import pandas as pd
import datetime
from scrapy.xlib.pydispatch import dispatcher

from jobcrawl.mailer import send_email
from jobcrawl.clientchanges import ClientChanges
from excel_gen import generate_excel

today = datetime.date.today()
today_str = today.strftime("%Y_%m_%d")


class LeftCompany(scrapy.Spider):
    """ Spider to verify removed companies """
    name = "left"
    allowed_domains = ["jobmaster.co.il", "drushim.co.il", "alljobs.co.il"]

    start_urls = []
    excel_dir = 'daily_competitor_client_changes'
    excel_path = '{}/{}_Daily-Competitor-Client-Change.xlsx'.format(
        excel_dir, today_str)

    def __init__(self):

        dispatcher.connect(self.spider_closed, signals.spider_closed)
        sys.stdout = codecs.getwriter(
            locale.getpreferredencoding())(sys.stdout)
        reload(sys)
        sys.setdefaultencoding('utf-8')

        # prepare clientchanges report
        self.c = ClientChanges()
        self.c.start()

        self.wb = load_workbook(self.excel_path)
        self.left_sheet = self.wb.get_sheet_by_name('Companies_That_left')
        self.new_sheet = self.wb.get_sheet_by_name('New_Companies')
        self.new_sheet_write = self.wb.create_sheet('New', 0)
        self.left_sheet_write = self.wb.create_sheet('Left', 1)

    def start_requests(self):
        wb = open_workbook(self.excel_path)
        sheet = wb.sheet_by_name('Companies_That_left')

        # left_header_row = sheet.row_values(0)
        # font = Font(size=11, bold=True)
        # self.left_sheet_write.append(left_header_row)
        # for i in ['A', 'B', 'C', 'D']:
        #     c = self.left_sheet_write[i + '1']
        #     c.font = font

        for i in range(1, sheet.nrows):
            row = sheet.row_values(i)
            company = row[1]
            site = row[0]
            company_url = row[2]
            company_jobs = row[3]
            if company_url:
                company_detail = [site, company, company_url, company_jobs]
                yield scrapy.Request(
                    company_url, self.parse,
                    meta={'company_detail': company_detail, 'type': 'removed'},
                    dont_filter=True
                )
        # get links for new companies
        new_sheet = wb.sheet_by_name('New_Companies')

        new_header_row = new_sheet.row_values(0)
        font = Font(size=11, bold=True)
        self.new_sheet_write.append(new_header_row)
        for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            c = self.new_sheet_write[i + '1']
            c.font = font

        for i in range(1, new_sheet.nrows):
            new_row = new_sheet.row_values(i)
            new_company = new_row[1]
            new_site = new_row[0]
            new_company_url = new_row[2]
            new_company_jobs = new_row[3]
            if new_company_url:
                new_company_detail = [
                    new_site, new_company, new_company_url, new_company_jobs]
                yield scrapy.Request(
                    new_company_url, self.parse,
                    meta={'company_detail': new_company_detail, 'type': 'new'},
                    dont_filter=True
                )

    def parse(self, response):

        company_detail = response.meta['company_detail']
        if response.meta['type'] == 'removed':
            drushimob_div = response.xpath("//div[@class='jobCount']")
            alljobs_jobs_div = response.xpath("//div[@class='job-paging']")
            jobmaster_jobs = response.xpath(
                "//div[@class='CenterContent']/article")

            if (
                drushimob_div or alljobs_jobs_div or jobmaster_jobs or
                '/Search/' in response.url
            ):
                # wb = load_workbook(self.excel_path)
                # sheet = wb.get_sheet_by_name('Companies_That_left')

                self.left_sheet.append(company_detail)
                # self.wb.save(self.excel_path)
        else:
            # open alljobs
            company_site_url = ''
            if company_detail[0] == 'AllJobs':
                try:
                    company_site_url = response.xpath(
                        "//div[@id='divTagCompanyCategory']/"
                        "following-sibling::div"
                    )[0].xpath("./a/text()").extract_first()
                except:
                    pass

            company_detail.extend([company_site_url, '', ''])
            self.new_sheet_write.append(company_detail)

    def spider_closed(self, spider):
        self.wb.remove_sheet(self.new_sheet)
        self.new_sheet_write.title = "New_Companies"
        self.wb.save(self.excel_path)

        new_companies_df = pd.read_excel(
            self.excel_path, sheetname='New_Companies')
        new_companies_df = new_companies_df.sort_values(
            by=['Site', 'Company'])
        left_companies_df = pd.read_excel(
            self.excel_path, sheetname='Companies_That_left')
        left_companies_df = left_companies_df.drop_duplicates(keep=False)
        left_companies_df = left_companies_df.sort_values(
            by=['Site', 'Company'])

        writer = pd.ExcelWriter(self.excel_path, engine='openpyxl')
        new_companies_df.to_excel(writer, 'New_Companies', index=False)
        left_companies_df.to_excel(writer, 'Companies_That_left', index=False)
        writer.save()

        # send email for competitior changes
        directory = 'daily_competitor_client_changes'
        file_name = '{}_Daily-Competitor-Client-Change.xlsx'.format(
            today_str)

        self.stats = self.c.get_stats()

        body = """
            Please find the attachment for {subject}.

            --- New / Removed Companies per Site ---
            Drushim : (new) {drushim_new}, (removed) {drushim_removed}
            JobMaster : (new) {jobmaster_new}, (removed) {jobmaster_removed}
            AllJobs : (new) {alljobs_new}, (removed) {alljobs_removed}

            --- New Companies ---
            Drushim : {drushim_new}
            JobMaster : {jobmaster_new}
            AllJobs : {alljobs_new}

            --- Removed Companies ---
            Drushim : {drushim_removed}
            JobMaster : {jobmaster_removed}
            AllJobs : {alljobs_removed}
        """.format(
            subject=file_name, drushim_new=self.stats['new']['Drushim'],
            drushim_removed=self.stats['removed']['Drushim'],
            jobmaster_new=self.stats['new']['JobMaster'],
            jobmaster_removed=self.stats['removed']['JobMaster'],
            alljobs_new=self.stats['new']['AllJobs'],
            alljobs_removed=self.stats['removed']['AllJobs']
        )

        send_email(directory=directory, file_name=file_name, body=body)

        # send an email for 3 excel attachments
        directory = "IL-jobcrawl-data"
        file_to_send = []
        for site in ['Drushim', 'Alljobs', 'Jobmaster']:
            file_name = '{}_{}.xlsx'.format(
                today_str, site)
            # check if the file is corrupt
            try:
                load_workbook('{}/{}'.format(directory, file_name))
                print('{} File good'.format(site))
            except:
                print('{} file corrupt, regenerationg'.format(site))
                # file is corrupt, generate from sql
                generate_excel(site)
                print('{} File generation success'.format(site))
            file_to_send.append(file_name)

        subject = '{}_Daily-List-Of-Competitor-Jobs.xlsx'.format(
            file_to_send[0][:10])
        body = """
            Please find the attachment for {subject}.

            --- Jobs / Companies per Site ---
            Drushim : (jobs) {drushim_jobs}, (companies) {drushim_companies}
            JobMaster : (jobs) {jm_jobs}, (companies) {jm_companies}
            AllJobs : (jobs) {alljobs_jobs}, (companies) {alljobs_companies}

            --- Jobs per Site ---
            Drushim : {drushim_jobs} jobs
            JobMaster : {jm_jobs} jobs
            AllJobs : {alljobs_jobs} jobs

            --- Companies per Site ---
            Drushim : {drushim_companies} companies
            JobMaster : {jm_companies} companies
            AllJobs : {alljobs_companies} companies
        """.format(
            subject=subject, drushim_jobs=self.stats['total_jobs']['Drushim'],
            drushim_companies=self.stats['total_companies']['Drushim'],
            jm_jobs=self.stats['total_jobs']['JobMaster'],
            jm_companies=self.stats['total_companies']['JobMaster'],
            alljobs_jobs=self.stats['total_jobs']['AllJobs'],
            alljobs_companies=self.stats['total_companies']['AllJobs']
        )

        send_email(
            directory=directory, file_name=file_to_send, body=body, multi=True)
